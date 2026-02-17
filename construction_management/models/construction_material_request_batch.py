from odoo import models, fields, api, _

class ConstructionMaterialRequestBatch(models.Model):
    _name = 'construction.material.request.batch'
    _description = 'Material Request Batch'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = 'create_date desc'

    def action_export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            # We don't raise here to avoid crashing, but the caller should handle None
            return None

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "So‘rov"

        # --- Styles ---
        # Fonts
        font_title = Font(name='Arial', size=14, bold=True)
        font_info = Font(name='Arial', size=10)
        font_header = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        font_data = Font(name='Arial', size=10)
        font_total = Font(name='Arial', size=11, bold=True)
        
        # Borders
        thin = Side(border_style="thin", color="000000")
        border_all = Border(top=thin, left=thin, right=thin, bottom=thin)
        
        # Fills
        fill_header = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        # --- Report Info ---
        # Assume single project context if multiple batches
        project_name = self[0].project_id.name if self else ""
        client_name = self[0].project_id.customer_id.name or "Noma'lum"
        usta_name = self[0].requester_id.name or ""
        snab_name = self.env.user.name
        
        # Determine "Document Number" and "Date"
        if len(self) == 1:
            doc_number = self[0].name or str(self[0].id)
            doc_date = self[0].date.strftime('%d.%m.%Y')
        else:
            doc_number = f"Ro‘yxat ({len(self)} ta)"
            doc_date = fields.Date.today().strftime('%d.%m.%Y')

        # Title
        ws.merge_cells('A1:I2')
        cell_title = ws['A1']
        cell_title.value = "Material so‘rovi"
        cell_title.font = font_title
        cell_title.alignment = Alignment(horizontal='center', vertical='center')

        # Info Block
        # Row 3
        ws.merge_cells('A3:D3')
        ws['A3'] = f"Loyiha: {project_name}"
        ws['A3'].font = font_info
        
        ws.merge_cells('E3:I3')
        ws['E3'] = f"Buyurtmachi: {client_name}"
        ws['E3'].font = font_info

        # Row 4
        ws.merge_cells('A4:D4')
        ws['A4'] = f"Usta (Мастер): {usta_name}"
        ws['A4'].font = font_info

        ws.merge_cells('E4:I4')
        ws['E4'] = f"Snab (Исполнитель): {snab_name}"
        ws['E4'].font = font_info
        
        # Row 5
        ws.merge_cells('A5:D5')
        ws['A5'] = f"So‘rov sanasi: {doc_date}"
        ws['A5'].font = font_info
        
        ws.merge_cells('E5:I5')
        ws['E5'] = f"Hujjat raqami: {doc_number}"
        ws['E5'].font = font_info

        # --- Table Headers ---
        headers = ["Наименование", "Ед-измер", "Кол-во", "Цена", "Сумма", "Дата", "Мастер", "Комментарий", "Исполнитель"]
        header_row_idx = 7
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row_idx, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.border = border_all
            cell.alignment = Alignment(horizontal='center')

        # --- Table Data ---
        row_idx = 8
        total_qty = 0
        total_sum = 0
        has_zero_price = False

        for batch in self:
            b_date = batch.date.strftime('%d.%m.%Y')
            b_master = batch.requester_id.name or ""
            b_comment = batch.task_id.name or batch.name or ""
            
            for line in batch.line_ids:
                qty = line.quantity
                price = line.unit_price or 0
                line_sum = qty * price if price > 0 else 0
                
                total_qty += qty
                total_sum += line_sum
                if price == 0:
                    has_zero_price = True

                # Row values
                vals = [
                    line.product_name,
                    "", # unit placeholder
                    qty,
                    price,
                    line_sum,
                    b_date,
                    b_master,
                    b_comment,
                    snab_name
                ]

                # Write Row
                for col_idx, val in enumerate(vals, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = font_data
                    cell.border = border_all
                    
                    if col_idx == 3: # Qty
                        cell.number_format = '0.00'
                    elif col_idx in [4, 5]: # Price, Sum
                        cell.number_format = '#,##0 "so\'m"'

                row_idx += 1

        # --- Totals ---
        row_idx += 1 # Skip one line? OR attached. attached is better.
        # Actually user said "After table...".
        
        # Total Qty
        cell_tq_label = ws.cell(row=row_idx, column=2, value="Итого Кол-во:")
        cell_tq_label.font = font_total
        cell_tq_label.alignment = Alignment(horizontal='right')
        
        cell_tq_val = ws.cell(row=row_idx, column=3, value=total_qty)
        cell_tq_val.font = font_total
        cell_tq_val.border = Border(top=thin)
        cell_tq_val.number_format = '0.00'

        # Total Sum
        cell_ts_label = ws.cell(row=row_idx, column=4, value="Итого Сумма:")
        cell_ts_label.font = font_total
        cell_ts_label.alignment = Alignment(horizontal='right')

        cell_ts_val = ws.cell(row=row_idx, column=5, value=total_sum)
        cell_ts_val.font = font_total
        cell_ts_val.border = Border(top=thin)
        cell_ts_val.number_format = '#,##0 "so\'m"'

        # Warning
        if has_zero_price:
            row_idx += 2
            ws.merge_cells(f'A{row_idx}:I{row_idx}')
            cell_warn = ws[f'A{row_idx}']
            cell_warn.value = "⚠️ Ba’zi pozitsiyalarda narx kiritilmagan (Цена=0)."
            cell_warn.font = Font(color="FF0000", italic=True)

        # Auto-width
        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_len = 0
            for cell in col_cells:
                if cell.value:
                    # simplistic width calc
                    try:
                        max_len = max(max_len, len(str(cell.value)))
                    except: pass
            adjusted_width = max_len + 2
            ws.column_dimensions[get_column_letter(col_idx)].width = min(adjusted_width, 50) # Cap width

        # Save
        import io
        import base64
        fp = io.BytesIO()
        wb.save(fp)
        fp.seek(0)
        datas = base64.b64encode(fp.read())
        
        name_prefix = self[0].name if len(self)==1 else "Request_List"
        att_name = f"{name_prefix}.xlsx"
        attachment = self.env['ir.attachment'].create({
            'name': att_name,
            'datas': datas,
            'res_model': self._name,
            'res_id': self[0].id if self else 0,
            'type': 'binary'
        })
        return attachment

    def action_export_pdf(self):
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.pdfbase import pdfmetrics
            from reportlab.lib.units import cm
        except ImportError:
            return None

        import io
        import base64
        from datetime import datetime

        # Register Font (FIX Cyrillic)
        try:
            from odoo.modules import get_module_resource
            font_path = get_module_resource('construction_management', 'static', 'fonts', 'DejaVuSans.ttf')
            font_bold_path = get_module_resource('construction_management', 'static', 'fonts', 'DejaVuSans-Bold.ttf')
            
            if font_path:
                pdfmetrics.registerFont(TTFont('DejaVuSans', font_path))
            if font_bold_path:
                pdfmetrics.registerFont(TTFont('DejaVuSans-Bold', font_bold_path))
        except Exception as e:
             _logger.warning(f"Font registration failed: {e}")
             pass

        fp = io.BytesIO()
        doc = SimpleDocTemplate(fp, pagesize=landscape(A4), topMargin=1*cm, bottomMargin=1*cm)
        elements = []
        styles = getSampleStyleSheet()
        
        # Custom Styles
        style_title = ParagraphStyle('MyTitle', parent=styles['Normal'], fontName='DejaVuSans-Bold', fontSize=16, alignment=1, spaceAfter=20)
        style_normal = ParagraphStyle('MyDesc', parent=styles['Normal'], fontName='DejaVuSans', fontSize=10, leading=14)
        style_footer = ParagraphStyle('MyFooter', parent=styles['Normal'], fontName='DejaVuSans', fontSize=8, textColor=colors.grey)

        # Helper to format money
        def fmt_money(val):
            return "{:,.0f} UZS".format(val).replace(',', ' ')

        # --- Report Info ---
        project_name = self[0].project_id.name if self else ""
        client_name = self[0].project_id.customer_id.name or "Noma'lum"
        usta_name = self[0].requester_id.name or ""
        snab_name = self.env.user.name
        
        if len(self) == 1:
            doc_number = self[0].name or str(self[0].id)
            doc_date = self[0].date.strftime('%d.%m.%Y')
        else:
            doc_number = f"Ro‘yxat ({len(self)} ta)"
            doc_date = fields.Date.today().strftime('%d.%m.%Y')

        # Title
        elements.append(Paragraph("Material so‘rovi", style_title))

        # Info Block (2 cols via Table)
        info_data = [
            [f"Loyiha: {project_name}", f"Buyurtmachi: {client_name}"],
            [f"Usta (Мастер): {usta_name}", f"Snab (Исполнитель): {snab_name}"],
            [f"So‘rov sanasi: {doc_date}", f"Hujjat raqami: {doc_number}"]
        ]
        info_table = Table(info_data, colWidths=[120, 150]) # adjust columns
        # Actually use Paragraphs inside table to wrap text? Or just simple strings.
        # Strings better for simple info.
        # But we need Cyrillic support, so font is critical here.
        # If Table cells are plain strings, reportlab uses default font?
        # No, we can style the table.
        
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0,0), (-1,-1), 'DejaVuSans'),
            ('FONTSIZE', (0,0), (-1,-1), 10),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ]))
        # Improve layout: spread across page width
        # A4 Landscape is ~297mm wide.
        # Let's use 50% 50%.
        avail_width = doc.width
        info_table = Table(info_data, colWidths=[avail_width/2, avail_width/2])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0,0), (-1,-1), 'DejaVuSans'),
            ('FONTSIZE', (0,0), (-1,-1), 10),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ]))
        
        elements.append(info_table)
        elements.append(Spacer(1, 10))

        # --- Main Table ---
        headers = ["Наименование", "Ед", "Кол-во", "Цена", "Сумма", "Дата", "Мастер", "Коммент", "Исполн."] 
        # Shortened a bit for fit if needed
        
        data = [headers]
        
        total_qty = 0
        total_sum = 0
        
        for batch in self:
            b_date = batch.date.strftime('%d.%m.%Y')
            b_master = batch.requester_id.name or ""
            b_comment = batch.task_id.name or batch.name or ""
            
            for line in batch.line_ids:
                qty = line.quantity
                price = line.unit_price or 0
                line_sum = qty * price if price > 0 else 0
                
                total_qty += qty
                total_sum += line_sum
                
                row = [
                    line.product_name,
                    "",
                    f"{qty:g}",
                    fmt_money(price) if price > 0 else "0",
                    fmt_money(line_sum),
                    b_date,
                    b_master,
                    b_comment[:20], # Truncate comment?
                    snab_name
                ]
                data.append(row)
        
        # Totals Row
        data.append(["", "Итого:", f"{total_qty:g}", "", fmt_money(total_sum), "", "", "", ""])

        # Table Layout
        # Auto calculate col widths? Or fixed.
        # Cols: Name(30%), Ed(5%), Qty(8%), Price(12%), Sum(12%), Date(8%), Master(10%), Comm(10%), Isp(5%)
        w = doc.width
        cw = [w*0.25, w*0.05, w*0.08, w*0.12, w*0.12, w*0.08, w*0.10, w*0.10, w*0.10]
        
        t = Table(data, colWidths=cw, repeatRows=1)
        
        # Styles
        ts = [
            ('FONTNAME', (0,0), (-1,-1), 'DejaVuSans'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
            ('FONTNAME', (0,0), (-1,0), 'DejaVuSans-Bold'),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('ALIGN', (0,1), (0,-1), 'LEFT'), # Name left
            ('ALIGN', (7,1), (7,-1), 'LEFT'), # Comment left
        ]
        # Totals Row Style (Last row)
        ts.append(('FONTNAME', (0,-1), (-1,-1), 'DejaVuSans-Bold'))
        ts.append(('BACKGROUND', (0,-1), (-1,-1), colors.whitesmoke))
        
        t.setStyle(TableStyle(ts))
        elements.append(t)
        
        elements.append(Spacer(1, 20))
        
        # Footer
        now_str = datetime.now().strftime('%d.%m.%Y %H:%M')
        elements.append(Paragraph(f"Yaratilgan sana: {now_str}", style_footer))
        
        doc.build(elements)
        
        fp.seek(0)
        datas = base64.b64encode(fp.read())
        
        name_prefix = self[0].name if len(self)==1 else "Request_List"
        att_name = f"{name_prefix}.pdf"
        attachment = self.env['ir.attachment'].create({
            'name': att_name,
            'datas': datas,
            'res_model': self._name,
            'res_id': self[0].id if self else 0,
            'type': 'binary'
        })
        return attachment

    name = fields.Char(string='So‘rov raqami', required=True, copy=False, readonly=True, index=True, default=lambda self: _('Yangi'))
    
    project_id = fields.Many2one('construction.project', string='Loyiha', required=True, tracking=True)
    requester_id = fields.Many2one('res.users', string='Usta', readonly=True, tracking=True)
    task_id = fields.Many2one('construction.work.task', string='Vazifa', readonly=True, tracking=True)
    date = fields.Date(string='Sana', default=fields.Date.context_today)
    
    state = fields.Selection([
        ('draft', 'Yuborildi'),
        ('priced', 'Narx qo‘yildi'),
        ('approved', 'Tasdiqlandi'),
        ('rejected', 'Rad etildi')
    ], string='Holat', default='draft', tracking=True)
    
    line_ids = fields.One2many('construction.material.request.line', 'batch_id', string='Qatorlar')
    
    approve_user_id = fields.Many2one('res.users', string='Tasdiqlagan', readonly=True)
    approve_date = fields.Datetime(string='Tasdiqlangan sana', readonly=True)

    delivery_count = fields.Integer(compute='_compute_delivery_count', string='Yetkazib berish')
    delivery_state = fields.Selection([
        ('purchased', 'Sotib olindi'),
        ('in_transit', 'Yo‘lda'),
        ('delivered', 'Yetkazildi')
    ], string='Yetkazib berish holati', compute='_compute_delivery_info')
    delivery_updated_at = fields.Datetime(string='Yangilangan vaqt', compute='_compute_delivery_info')

    @api.depends('delivery_count') # delivery_count logic searches DB, so we can piggyback or just depend on nothing/write
    def _compute_delivery_info(self):
        for rec in self:
            delivery = self.env['construction.material.delivery'].search([('batch_id', '=', rec.id)], limit=1)
            rec.delivery_state = delivery.state if delivery else False
            rec.delivery_updated_at = delivery.updated_at if delivery else False

    def _compute_delivery_count(self):
        for rec in self:
            rec.delivery_count = self.env['construction.material.delivery'].search_count([('batch_id', '=', rec.id)])

    def action_open_delivery(self):
        self.ensure_one()
        delivery = self.env['construction.material.delivery'].search([('batch_id', '=', self.id)], limit=1)
        if not delivery:
            delivery = self.env['construction.material.delivery'].create({'batch_id': self.id})
            
        return {
            'type': 'ir.actions.act_window',
            'name': 'Yetkazib berish holati',
            'res_model': 'construction.material.delivery',
            'res_id': delivery.id,
            'view_mode': 'form',
            'target': 'current'
        }
    
    @api.model
    def create(self, vals):
        if vals.get('name', _('Yangi')) == _('Yangi'):
            vals['name'] = self.env['ir.sequence'].next_by_code('construction.material.request.batch') or _('Yangi')
        return super(ConstructionMaterialRequestBatch, self).create(vals)

    def action_reset_to_draft(self):
        # Reset approval info
        self.write({
            'approve_user_id': False,
            'approve_date': False
        })
        
        # Check if any line has price
        has_price = any(line.unit_price > 0 for line in self.line_ids)
        
        if has_price:
            self.write({'state': 'priced'})
            # Notify via bot if available
            if 'construction.telegram.bot' in self.env:
                self.env['construction.telegram.bot']._system_send_batch_approval(self)
        else:
            self.write({'state': 'draft'})
            # Notify via bot if available
            if 'construction.telegram.bot' in self.env:
                self.env['construction.telegram.bot']._system_notify_snab_new_batch(self)


class ConstructionMaterialRequestLine(models.Model):
    _name = 'construction.material.request.line'
    _description = 'Material Request Line'

    batch_id = fields.Many2one('construction.material.request.batch', string='So‘rov', required=True, ondelete='cascade')
    product_name = fields.Char(string='Mahsulot nomi', required=True)
    quantity = fields.Float(string='Miqdor', required=True)
    
    unit_price = fields.Float(string='Birlik narxi (so‘m)')
    total_price = fields.Float(string='Jami (so‘m)', compute='_compute_total_price', store=True)

    @api.depends('quantity', 'unit_price')
    def _compute_total_price(self):
        for line in self:
            line.total_price = line.quantity * line.unit_price
